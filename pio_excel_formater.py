from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import re
import win32com.client
import csv
import os.path

from time import time

#Defining various styling templates that will be used for formatting cells of the "data" document

alignment_settings = Alignment(horizontal="center", vertical="center")

border_settings1 = Border(left=Side(border_style='thin'),
							right=Side(border_style='thin'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='thin'))

border_settings2 = Border(left=Side(border_style='thick'),
							right=Side(border_style='thick'),
							top=Side(border_style='thick'),
							bottom=Side(border_style='thick'))

border_settings3 = Border(left=Side(border_style='medium'),
							right=Side(border_style='medium'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='medium'))

border_settings4 = Border(left=Side(border_style='medium'),
							right=Side(border_style='thin'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='thin'))

border_settings5 = Border(left=Side(border_style='thin'),
							right=Side(border_style='medium'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='thin'))

border_settings6 = Border(left=Side(border_style='thin'),
							right=Side(border_style='thin'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='thin'))

border_settings7 = Border(left=Side(border_style='thin'),
							right=Side(border_style='thin'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='medium'))

def initial_load():
	global wb
	wb = load_workbook('data.xlsx')

def save_document_excel():
	wb.save('data.xlsx')


def update_excel_sheet(name, JMBG, address, city, credit, instalments, instalment_info, receipt, store):

	'''

	Editing the "Zbirno" main sheet of the 'data.xlsx' document

	1. Setting up the specific index value that will be assigned to each new customer
		--(no. of the last populated row -6. We used -6 because the rows start at the 8th row) 
	
	2. Setting up the row_counter value to detect the last populated row in the sheet. row_counter will be used to append each new row with new customer
		and it's personal data

	3. Zipping two lists. One is with column names, the other one is containing data entered in the program inside of entry boxes.

	4. Setting up the thousands separator for each new cell in G and I columns. This is where credit value and instalment value will be appended
	
	'''
	ws = wb["Zbirno"]	

	index = len(ws['A']) - 6

	row_counter = str(len(ws["B"])+1)

	instalment_info = instalment_info.get().replace(",","")

	credit = credit.get().replace(",","")

	instalments = instalments.get().replace(",","")

	entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
	entry_data = [index, name.get(), address.get(), city.get(), JMBG.get(), float(credit), int(instalments), float(instalment_info), receipt.get(), store.get()]

	for col, data in zip(entry_data_cell, entry_data):
		ws[col + row_counter] = data
		ws[col + row_counter].alignment = alignment_settings
		ws[col + row_counter].border = border_settings1

	ws["G" + row_counter].number_format = '#,##0.00'
	ws["I" + row_counter].number_format = '#,##0.00'

	'''
	Creating and editing a new sheet for each new customer appended to the database.

	1. Appending customer-unique index number and concatenating it to the customers name

	2. Creating a new sheet for customer using that specific string name

	3. Setting the width of each column inside customer-specific sheet.

	4. Setting zoom attribute for each sheet

	'''

	new_sheet_name = str(index) + ". " + name.get()
	wb.create_sheet(title= new_sheet_name)

	ws1 = wb[new_sheet_name]
	sheet_ranges = wb[new_sheet_name]
    
	sheet_ranges.column_dimensions["A"].width = 15
	sheet_ranges.column_dimensions["B"].width = 15
	sheet_ranges.column_dimensions["C"].width = 15
	sheet_ranges.column_dimensions["D"].width = 15
	sheet_ranges.column_dimensions["E"].width = 15
	sheet_ranges.column_dimensions["F"].width = 15
	sheet_ranges.column_dimensions["G"].width = 15
	sheet_ranges.column_dimensions["H"].width = 15
	sheet_ranges.column_dimensions["I"].width = 15
	sheet_ranges.column_dimensions["J"].width = 15
	sheet_ranges.column_dimensions["K"].width = 15
	sheet_ranges.column_dimensions["L"].width = 15
	sheet_ranges.column_dimensions["M"].width = 15
	sheet_ranges.column_dimensions["N"].width = 15

	sheet_ranges.row_dimensions[16].height = 63
	sheet_ranges.row_dimensions[17].height = 63
	sheet_ranges.row_dimensions[18].height = 63

    
	ws1.sheet_view.zoomScale = 70

	'''
	Editing cells inside of the sheet to accommodate data user has entered inside of the program - part 1. 

	1. Merging columns C to L where the user name will be used as a title of a user-specific sheet

	2. Merging cells to create a square container where "maloprodaja" will be stored

	3. Merging another several rows of cells to create a table-like data container where another copy of user-related info will be stored

	4. Merging cells to create a container that will indicate whether the credit has been paid of or not. 
	Also this is where the info of user passing away will be stored

	'''

	
	ws1['C1'] = new_sheet_name
	ws1['C1'].font = Font(size=36, bold=True)
	ws1['C1'].border = border_settings2
	ws1['C1'].alignment = alignment_settings
	ws1.merge_cells('C1:L3')


	ws1['L6'] = store.get()
	ws1['L6'].font = Font(size=36, bold=True)
	ws1['L6'].border = border_settings3
	ws1['L6'].alignment = alignment_settings
	ws1.merge_cells('L6:L9')


	ws1['C6'] = "Grad"
	ws1['C6'].font = Font(size=16)
	ws1['C6'].border = border_settings3
	ws1['C6'].alignment = alignment_settings
	ws1.merge_cells('C6:D7')


	ws1['C8'] = "Adresa"
	ws1['C8'].font = Font(size=16)
	ws1['C8'].border = border_settings3
	ws1['C8'].alignment = alignment_settings
	ws1.merge_cells('C8:D9')


	ws1['C10'] = "JMBG"
	ws1['C10'].font = Font(size=16)
	ws1['C10'].border = border_settings3
	ws1['C10'].alignment = alignment_settings
	ws1.merge_cells('C10:D11')	


	ws1['C12'] = "Broj računa"
	ws1['C12'].font = Font(size=16)
	ws1['C12'].border = border_settings3
	ws1['C12'].alignment = alignment_settings
	ws1.merge_cells('C12:D13')


	ws1['E6'] = city.get()
	ws1['E6'].font = Font(size=16)
	ws1['E6'].border = border_settings3
	ws1['E6'].alignment = alignment_settings
	ws1.merge_cells('E6:G7')


	ws1['E8'] = address.get()
	ws1['E8'].font = Font(size=16)
	ws1['E8'].border = border_settings3
	ws1['E8'].alignment = alignment_settings
	ws1.merge_cells('E8:G9')


	ws1['E10'] = JMBG.get()
	ws1['E10'].font = Font(size=16)
	ws1['E10'].border = border_settings3
	ws1['E10'].alignment = alignment_settings
	ws1.merge_cells('E10:G11')


	ws1['E12'] = receipt.get()
	ws1['E12'].font = Font(size=16)
	ws1['E12'].border = border_settings3
	ws1['E12'].alignment = alignment_settings
	ws1.merge_cells('E12:G13')

	ws1['I12'] = "Otplata u toku"
	ws1['I12'].font = Font(size=16)
	ws1['I12'].border = border_settings3
	ws1['I12'].alignment = alignment_settings
	ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")
	ws1.merge_cells('I12:L13')



	'''
	Editing cells inside of the sheet to accommodate data user has entered inside of the program - part 2.
	
	1. Defining table of instalments header by merging header columns. 
	We start with "Pocetno", then we add roman numerals  columns based on the number of instalments. 
	Finally we end our list with "Ostatak" column

	2. We zip a list of 


	'''

	header_list = ["Početno"]

	instalments_list = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]

	end_list = "Ostatak"

	for no in range(int(instalments)):
		header_list.append(instalments_list[no])
    
	header_list.append(end_list)
	
	for col, column_name in zip(list(range(1, len(header_list)+1)), header_list):
		c_letter = get_column_letter(col)  
		ws1[c_letter + "16"] = column_name
		ws1[c_letter + "16"].border = border_settings6
		ws1[c_letter + "16"].alignment = alignment_settings
		ws1[c_letter + "16"].font = Font(size=20, bold=True)

	for col in range(2,len(header_list)+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].border = border_settings1
		ws1[c_letter + "17"].alignment = alignment_settings
		ws1[c_letter + "17"].font = Font(size=16)
		ws1[c_letter + "17"].number_format = '#,##0.00'

	for col in range(2,len(header_list)):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].value = 0
		ws1[c_letter + "17"].fill = PatternFill("solid", fgColor="00FFCC99")

	for col in range(1,len(header_list)+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "18"].border = border_settings7
		ws1[c_letter + "18"].alignment = alignment_settings
		ws1[c_letter + "18"].font = Font(size=16)

	for col in range(2,len(header_list)):
		c_letter = get_column_letter(col)
		ws1[c_letter + "18"].fill = PatternFill("solid", fgColor="00FFCC99")

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[len(header_list)]

	ws1['A16'].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='thin'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='thin'))
	ws1['A16'].alignment = alignment_settings	
	
	ws1["A17"] = float(credit)
	ws1["A17"].number_format = '#,##0.00'

	ws1['A17'].border = border_settings4
	ws1['A17'].alignment = alignment_settings
	ws1['A17'].font = Font(size=16)		
		

	ws1["A18"] = "Datum Uplate"
	ws1['A18'].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='thin'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='medium'))
	ws1['A18'].alignment = alignment_settings
	ws1['A18'].font = Font(size=12, bold=True)

	
	ws1[last_row + '16'].border = Border(left=Side(border_style='thin'),
							right=Side(border_style='medium'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='thin'))
	ws1[last_row + '16'].alignment = alignment_settings	


	ws1[last_row + '17'].border = border_settings5
	ws1[last_row + '17'].alignment = alignment_settings

	middle_rows = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=len(header_list)-1):
		for cell in row:
			cell_name = cell.coordinate
			cell_value = ws1[cell_name].value
			middle_rows.append(cell_value)

	

	
	ws1[ last_row + "17"].number_format = '#,##0.00'
	ws1[ last_row + "17"] = ws1["A17"].value - sum(middle_rows)
	ws1[ last_row + "17"].value = round(float(ws1[ last_row + "17"].value),2)	
	
	ws1[last_row + '18'].border = Border(left=Side(border_style='thin'),
							right=Side(border_style='medium'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='medium'))
	ws1[last_row + '18'].alignment = alignment_settings

	ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
	ws1.sheet_properties.pageSetUpPr.fitToPage = True
	ws1.page_setup.fitToHeight = False

	
	ws1['A20'] = "Napomene"
	ws1['A20'].font = Font(size=26)
	ws1['A20'].border = border_settings3
	ws1['A20'].alignment = alignment_settings
	ws1.merge_cells('A20:N21')

	
	ws1['A22'].border = border_settings3
	ws1['A22'].value = ""
	ws1['A22'].font = Font(size=16)
	ws1['A22'].alignment = Alignment(horizontal="left", vertical="top")
	ws1.merge_cells('A22:N30')


	wb.save('data.xlsx')		


def create_excel_sheet():
	wb = Workbook()

	ws = wb.active

	ws.title = "Zbirno"

	sheet_ranges = wb["Zbirno"]
	sheet_ranges.column_dimensions["B"].width = 8
	sheet_ranges.column_dimensions["C"].width = 25
	sheet_ranges.column_dimensions["D"].width = 25
	sheet_ranges.column_dimensions["E"].width = 20
	sheet_ranges.column_dimensions["F"].width = 15
	sheet_ranges.column_dimensions["G"].width = 15
	sheet_ranges.column_dimensions["H"].width = 8
	sheet_ranges.column_dimensions["I"].width = 15
	sheet_ranges.column_dimensions["J"].width = 15
	sheet_ranges.column_dimensions["K"].width = 5


	ws['B2'] = "PIO EVIDENCIJA 2022"
	ws['B2'].font = Font(size=36, bold=True)
	ws['B2'].border = Border(left=Side(border_style='thick'),
										right=Side(border_style='thick'),
										top=Side(border_style='thick'),
										bottom=Side(border_style='thick'))
	ws['B2'].alignment = Alignment(horizontal='center', vertical= 'center')
			
	ws.merge_cells('B2:K3') 

	for col, column_name in zip(list(range(2,12)), ["Indeks", "Ime i prezime", "Adresa", "Grad", "JMBG", "Iznos kredita", "Br. rata", "Iznos rate", "Broj računa", "MP"]):
		c_letter = get_column_letter(col)  
		ws[c_letter + "7"] = column_name
		ws[c_letter + "7"].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='medium'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='medium'))
		ws[c_letter + "7"].alignment = Alignment(horizontal='center', vertical= 'center')

	wb.save('data.xlsx')

def create_unidentified_payments_excel_table():

	wb = Workbook()

	ws = wb.active

	ws.title = "Lista uplata"

	sheet_ranges = wb["Lista uplata"]
	sheet_ranges.column_dimensions["B"].width = 15
	sheet_ranges.column_dimensions["C"].width = 15
	sheet_ranges.column_dimensions["D"].width = 15
	sheet_ranges.column_dimensions["E"].width = 8
	sheet_ranges.column_dimensions["F"].width = 25
	sheet_ranges.column_dimensions["G"].width = 25
	sheet_ranges.column_dimensions["H"].width = 15
	sheet_ranges.column_dimensions["I"].width = 15
	sheet_ranges.column_dimensions["J"].width = 15
	sheet_ranges.column_dimensions["K"].width = 15
	

	ws['B2'] = "NEPROKNJIŽENE UPLATE 2022"
	ws['B2'].font = Font(size=36, bold=True)
	ws['B2'].border = Border(left=Side(border_style='thick'),
										right=Side(border_style='thick'),
										top=Side(border_style='thick'),
										bottom=Side(border_style='thick'))
	ws['B2'].alignment = Alignment(horizontal='center', vertical= 'center')
			
	ws.merge_cells('B2:K3')

	for col, column_name in zip(list(range(5,9)), ["Indeks", "Ime i prezime", "Iznos uplate", "Datum uplate"]):
		c_letter = get_column_letter(col)  
		ws[c_letter + "7"] = column_name
		ws[c_letter + "7"].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='medium'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='medium'))
		ws[c_letter + "7"].alignment = Alignment(horizontal='center', vertical= 'center')

	
	wb.save('unidentified_payments.xlsx')

def format_by_store(store_index):

	data = []

	indexes = []

	ws = wb["Zbirno"]


	for row in ws.iter_rows(min_row=8):
		for cell in row:
			if cell.value == store_index.get():
				name = row[2].value
				index = row[1].value
				data.append(name)
				indexes.append(index)

	return data, indexes


def format_by_late_users():

	ws = wb["Zbirno"]

	names = []

	indexes = []

	for row in ws.iter_rows(min_row=8):
		sheet_name = str(row[1].value) + ". " + row[2].value
		ws1 = wb[sheet_name]
		status = ws1["I12"].value
		if status == "Korisnik preminuo":
			names.append(row[2].value)
			indexes.append(row[1].value)

	return names, indexes
	

def populate_listbox(drop):

	ws = wb["Zbirno"]

	choice = drop.get()

	min_row=0
	min_col=0
	max_col=0

	names = []
	indexes = []


	if choice == "Odaberi filter:":
		pass

	else:

		if choice == "Ime i prezime":
			min_row = 8
			min_col = 3
			max_col = 3
		if choice == "Br. računa":
			min_row = 8
			min_col = 10
			max_col = 10
		if choice == "Iznos rate":
			min_row = 8
			min_col = 9
			max_col = 9

		for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
			for cell in row:
				names.append(cell.value)

		for row in ws.iter_rows(min_row=8, min_col=2, max_col=2):
			for cell in row:
				indexes.append(cell.value)

	return names, indexes


def confirm_selection_excel(user_index):

	balance = 0

	ws = wb["Zbirno"]

	data = []

	instalment_data = []

	date_data = []

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
			data.extend([row[1].value, row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value,row[10].value])
			sheet_name = str(row[1].value) + ". " + row[2].value
			instalment = row[7].value

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[instalment+2]


	ws1 = wb[sheet_name]
	balance = ws1[last_row + "17"].value
	status = ws1["I12"].value

	if status == "Otplata u toku":
		status_color = "coral1"
	if status == "Kredit otplaćen":
		status_color = "#1abc9c"
	if status == "Korisnik preminuo":
		status_color = "red4"

	comment = ws1["A22"].value

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=instalment+1):
		for cell in row:
			instalment_data.extend([float(cell.value)])

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=instalment+1):
		for cell in row:
			date_data.extend([cell.value])

	data.extend([balance, status, status_color])

	return data, instalment_data, date_data, comment


def initial_unidentified_search(name):

	ws = wb["Zbirno"]

	names = []

	indexes = []

	for row in ws.iter_rows(min_row=8):
		if row[2].value == name:
			names.append(row[2].value)
			indexes.append(row[1].value)

	if len(names) == 0:
		condition = 0

	elif len(names) == 1:
		condition = 1
	else:
		condition = 2 

	return condition, indexes

def find_unidentified_user(name):

	balance = 0

	ws = wb["Zbirno"]

	data = []

	instalment_data = []

	date_data = []

	for row in ws.iter_rows(min_row=8):
		if row[2].value == name:
			data.extend([row[1].value, row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value,row[10].value])
			sheet_name = str(row[1].value) + ". " + row[2].value
			instalment = row[7].value

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[instalment+2]


	ws1 = wb[sheet_name]
	balance = ws1[last_row + "17"].value
	status = ws1["I12"].value

	if status == "Otplata u toku":
		status_color = "coral1"
	if status == "Kredit otplaćen":
		status_color = "#1abc9c"
	if status == "Korisnik preminuo":
		status_color = "red4"

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=instalment+1):
		for cell in row:
			instalment_data.extend([float(cell.value)])

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=instalment+1):
		for cell in row:
			date_data.extend([cell.value])

	data.extend([balance, status, status_color])

	return data, instalment_data, date_data

def edit_user_excel(user_index, name2, address2, city2, JMBG2, credit2, instalments2, instalment_info2, receipt2, store2):

	ws = wb["Zbirno"]

	data = []

	new_sheet_name = str(user_index) + ". " + name2.get()


	entry_data = [name2.get(), address2.get(), city2.get(), JMBG2.get(), float((credit2.get().replace(',',''))), int(instalments2.get()), float((instalment_info2.get().replace(',',''))), receipt2.get(), store2.get()]
	entry_data_cell = ["C", "D", "E", "F", "G", "H", "I", "J", "K"]
	

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
			old_instalment = row[7].value
			old_sheet_name = str(row[1].value) + ". " + row[2].value
			old_ws1 = wb[old_sheet_name]
			old_status = old_ws1["I12"].value
			for col, data in zip(entry_data_cell, entry_data):
				ws[col + str(user_index+7)] = data
			
	pattern = fr"^{user_index}\..+"

	for sheet in wb.worksheets:
		for match in re.finditer(pattern, sheet.title):
			ws1 = wb[match.group(0)]
			ws1['C1'] = new_sheet_name
			ws1.title = new_sheet_name
			comment = ws1["A22"].value

	ws1 = wb[new_sheet_name]

	instalment_data = []

	date_data = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=(old_instalment+1)):
		for cell in row:
			if cell.value !=0:
				instalment_data.extend([cell.value])

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=(old_instalment+1)):
		for cell in row:
			if cell.value !=None:
				date_data.extend([cell.value])

	ws1.delete_rows(16,15)

	
	ws1["L6"] = store2.get()
	ws1["E6"] = city2.get()
	ws1["E8"] = address2.get()
	ws1["E10"] = JMBG2.get()
	ws1["E12"] = receipt2.get()
	ws1["A17"] = float((credit2.get().replace(',','')))

	header_list = ["Početno"]

	instalments_list = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]

	end_list = "Ostatak"

	for no in range(int(instalments2.get())):
		header_list.append(instalments_list[no])
    
	header_list.append(end_list)
	
	for col, column_name in zip(list(range(1, len(header_list)+1)), header_list):
		c_letter = get_column_letter(col)  
		ws1[c_letter + "16"] = column_name
		ws1[c_letter + "16"].border = border_settings6
		ws1[c_letter + "16"].alignment = alignment_settings
		ws1[c_letter + "16"].font = Font(size=20, bold=True)

	for col in range(2,len(header_list)+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].border = border_settings1
		ws1[c_letter + "17"].alignment = alignment_settings
		ws1[c_letter + "17"].font = Font(size=16)
		ws1[c_letter + "17"].number_format = '#,##0.00'

	
	for col, data in zip(list(range(2,len(instalment_data)+2)), instalment_data):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"] = data	

		
	for col in range(2,len(header_list)):
		c_letter = get_column_letter(col)
		print(c_letter + "17")
		print( ws1[c_letter + "17"].value)
		if ws1[c_letter + "17"].value == None:
			ws1[c_letter + "17"].value = 0
	
	for col, data in zip(list(range(2,len(date_data)+2)), date_data):
		c_letter = get_column_letter(col)
		ws1[c_letter + "18"] = data


	for col in range(1,len(header_list)+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "18"].border = border_settings7
		ws1[c_letter + "18"].alignment = alignment_settings
		ws1[c_letter + "18"].font = Font(size=16)


	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[len(header_list)]

	ws1['A16'].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='thin'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='thin'))
	ws1['A16'].alignment = alignment_settings	
	

	ws1["A17"].number_format = '#,##0.00'
	ws1['A17'].border = border_settings4
	ws1['A17'].alignment = alignment_settings
	ws1['A17'].font = Font(size=16)		
		

	ws1["A18"] = "Datum Uplate"
	ws1['A18'].border = Border(left=Side(border_style='medium'),
							right=Side(border_style='thin'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='medium'))
	ws1['A18'].alignment = alignment_settings
	ws1['A18'].font = Font(size=12, bold=True)

	
	ws1[last_row + '16'].border = Border(left=Side(border_style='thin'),
							right=Side(border_style='medium'),
							top=Side(border_style='medium'),
							bottom=Side(border_style='thin'))
	ws1[last_row + '16'].alignment = alignment_settings	


	ws1[last_row + '17'].border = border_settings5
	ws1[last_row + '17'].alignment = alignment_settings

	middle_rows = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=len(header_list)-1):
		for cell in row:
			cell_name = cell.coordinate
			cell_value = ws1[cell_name].value
			middle_rows.append(cell_value)

	
	ws1[ last_row + "17"].number_format = '#,##0.00'
	ws1[ last_row + "17"] = ws1["A17"].value - sum(middle_rows)
	ws1[ last_row + "17"].value = round(float(ws1[ last_row + "17"].value),2)	
	
	ws1[last_row + '18'].border = Border(left=Side(border_style='thin'),
							right=Side(border_style='medium'),
							top=Side(border_style='thin'),
							bottom=Side(border_style='medium'))
	ws1[last_row + '18'].alignment = alignment_settings

	ws1['A20'] = "Napomene"
	ws1['A20'].font = Font(size=26)
	ws1['A20'].border = border_settings3
	ws1['A20'].alignment = alignment_settings
	ws1.merge_cells('A20:N21')

	
	ws1['A22'].border = border_settings3
	ws1['A22'].value = comment
	ws1['A22'].font = Font(size=16)
	ws1['A22'].alignment = Alignment(horizontal="left", vertical="top")
	ws1.merge_cells('A22:N30')

	if old_status == "Korisnik preminuo":
		ws1["I12"].value = "Korisnik preminuo"
		ws1['I12'].fill = PatternFill("solid", fgColor="BA1E0E")
		ws1['I12'].font = Font(color="FFFFFFFF", size=16)
	else:
		if ws1[last_row + "17"].value < 1:
			ws1["I12"].value = "Kredit otplaćen"
			ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
			ws1['I12'].font = Font(color="FF000000", size=16)

		else:
			ws1["I12"].value = "Otplata u toku"
			ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")
			ws1['I12'].font = Font(color="FF000000", size=16)

	table_header_length = len(header_list) - 2

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
				cell.font = Font(color="FFFFFFFF", size=16)
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
				cell.font = Font(color="FF000000", size=16)
			else:
				if cell.value !=0:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
					cell.font = Font(color="FF000000", size=16)
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")
					cell.font = Font(color="FF000000", size=16)

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
				cell.font = Font(color="FFFFFFFF", size=16)
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
				cell.font = Font(color="FF000000", size=16)
			else:
				if cell.value !=None:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
					cell.font = Font(color="FF000000", size=16)
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")
					cell.font = Font(color="FF000000", size=16)
		
def payment_excel(user_index, today):

	ws = wb["Zbirno"]

	cell_values = []

	cell_dates = []

	instalment_num = 0


	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
				sheet_name = str(row[1].value) + ". " + row[2].value
				instalment = row[8].value
				table_header_length = row[7].value


	ws1 = wb[sheet_name]

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[table_header_length+2]


	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_values.append(cell.value)

	for index, value in enumerate(cell_values):
		if value ==0:
			cell_values[index] = instalment
			break
	
	possible_entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

	entry_data_cell = []

	for no in range(table_header_length):
		entry_data_cell.append(possible_entry_data_cell[no])

	for item in cell_values:
		if item !=0:
			instalment_num += 1


	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for col, data in zip(entry_data_cell, cell_values):
			ws1[col + "17"] = data
			if ws1[col + "17"].value !=0:
				ws1[col + "17"].fill = PatternFill("solid", fgColor="A9F78D")
			else:
				ws1[col + "17"].fill = PatternFill("solid", fgColor="00FFCC99")

	for col in range(2,table_header_length+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].value = float(ws1[c_letter + "17"].value)
	
	middle_rows = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_name = cell.coordinate
			cell_value = ws1[cell_name].value
			middle_rows.append(cell_value)
	
	ws1[ last_row + "17"].number_format = '#,##0.00'
	ws1[ last_row + "17"] = ws1["A17"].value - sum(middle_rows)
	ws1[ last_row + "17"].value = round(float(ws1[ last_row + "17"].value),2)


	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_dates.append(cell.value)


	for index, date in enumerate(cell_dates):
		if date == None:
			cell_dates[index] = today
			break
			
	
	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for col, date in zip(entry_data_cell, cell_dates):
				ws1[col + "18"] = date
				
				if ws1[col + "18"].value !=None:
					ws1[col + "18"].fill = PatternFill("solid", fgColor="A9F78D")
				else:
					ws1[col + "18"].fill = PatternFill("solid", fgColor="00FFCC99")
				
	if ws1[last_row + "17"].value < 1:
		ws1["I12"].value = "Kredit otplaćen"
		ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
	
	else:
		ws1["I12"].value = "Otplata u toku"
		ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")


	return instalment_num

def atypical_payment_excel(user_index, instalment_entry, date_entry):

	ws = wb["Zbirno"]

	if type(instalment_entry) == int or type(instalment_entry) == float:
		instalment = instalment_entry

	else: 
		instalment = float((instalment_entry.get().replace(',','')))

	if type(date_entry) == str:
		today = date_entry

	else:
		today = date_entry.get()

	cell_values = []

	cell_dates = []

	instalment_num = 0

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
				sheet_name = str(row[1].value) + ". " + row[2].value
				table_header_length = row[7].value

	ws1 = wb[sheet_name]

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[table_header_length+2]


	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_values.append(cell.value)

	for index, value in enumerate(cell_values):
		if value ==0:
			cell_values[index] = instalment
			break

	possible_entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

	entry_data_cell = []

	for no in range(table_header_length):
		entry_data_cell.append(possible_entry_data_cell[no])

	for item in cell_values:
		if item !=0:
			instalment_num += 1
	

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for col, data in zip(entry_data_cell, cell_values):
			ws1[col + "17"] = data

	for col in range(2,table_header_length+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].value = float(ws1[c_letter + "17"].value)
	

	
	middle_rows = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_name = cell.coordinate
			cell_value = ws1[cell_name].value
			middle_rows.append(cell_value)
	
	ws1[ last_row + "17"].number_format = '#,##0.00'
	ws1[ last_row + "17"] = ws1["A17"].value - sum(middle_rows)
	ws1[ last_row + "17"].value = round(float(ws1[ last_row + "17"].value),2)


	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_dates.append(cell.value)
	

	for index, date in enumerate(cell_dates):
		if date == None:
			cell_dates[index] = today
			break
			
	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for col, date in zip(entry_data_cell, cell_dates):
				ws1[col + "18"] = date
	
				
	if ws1[last_row + "17"].value < 1:
		ws1["I12"].value = "Kredit otplaćen"
		ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
	
	else:
		ws1["I12"].value = "Otplata u toku"
		ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
			else:
				if cell.value !=0:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
			else:
				if cell.value !=None:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")
		

	return instalment_num


def cancel_payment_excel(user_index):

	ws = wb["Zbirno"]

	cell_values = []

	cell_dates = []

	instalment_num = 0

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
				sheet_name = str(row[1].value) + ". " + row[2].value
				table_header_length = row[7].value

	ws1 = wb[sheet_name]

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[table_header_length+2]

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_values.append(cell.value)
	
	cell_values.reverse()	

	for index, value in enumerate(cell_values):
		if value !=0:
			cell_values[index] = 0
			break
	
	cell_values.reverse()

	for item in cell_values:
		if item != 0:
			instalment_num += 1
	

	possible_entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

	entry_data_cell = []

	for no in range(table_header_length):
		entry_data_cell.append(possible_entry_data_cell[no])

	
	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for col, data in zip(entry_data_cell, cell_values):
			ws1[col + "17"] = data
			if ws1[col + "17"].value !=0:
				ws1[col + "17"].fill = PatternFill("solid", fgColor="A9F78D")
			else:
				ws1[col + "17"].fill = PatternFill("solid", fgColor="00FFCC99")

	for col in range(2,table_header_length+1):
		c_letter = get_column_letter(col)
		ws1[c_letter + "17"].value = float(ws1[c_letter + "17"].value)


	middle_rows = []

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_name = cell.coordinate
			cell_value = ws1[cell_name].value
			middle_rows.append(cell_value)
	
	ws1[ last_row + "17"].number_format = '#,##0.00'
	ws1[ last_row + "17"] = ws1["A17"].value - sum(middle_rows)
	ws1[ last_row + "17"].value = round(float(ws1[ last_row + "17"].value),2)


	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			cell_dates.append(cell.value)
	
	cell_dates.reverse()	

	for index, date in enumerate(cell_dates):
		if date !=None:
			cell_dates[index] = None
			break	

	cell_dates.reverse()

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for col, date in zip(entry_data_cell, cell_dates):
				ws1[col + "18"] = date
				if ws1[col + "18"].value !=None:
					ws1[col + "18"].fill = PatternFill("solid", fgColor="A9F78D")
				else:
					ws1[col + "18"].fill = PatternFill("solid", fgColor="00FFCC99")

	if ws1[last_row + "17"].value < 1:
		ws1["I12"].value = "Kredit otplaćen"
		ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
	
	else:
		ws1["I12"].value = "Otplata u toku"
		ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")

	return instalment_num

def late_user_excel(user_index, today):

	ws = wb["Zbirno"]

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
				sheet_name = str(row[1].value) + ". " + row[2].value
				table_header_length = row[7].value


	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	last_row = possible_cells[table_header_length+2]		

	ws1 = wb[sheet_name]

	possible_entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

	entry_data_cell = []

	for no in range(table_header_length):
		entry_data_cell.append(possible_entry_data_cell[no])


	if ws1["I12"].value != "Korisnik preminuo":
		ws1["I12"].value = "Korisnik preminuo"
		ws1['I12'].fill = PatternFill("solid", fgColor="BA1E0E")
		ws1['I12'].font = Font(color="FFFFFFFF", size=16)

	else:

		if ws1[last_row + "17"].value < 1:
			ws1["I12"].value = "Kredit otplaćen"
			ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
			ws1['I12'].font = Font(color="FF000000", size=16)
	
		else:
			ws1["I12"].value = "Otplata u toku"
			ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")
			ws1['I12'].font = Font(color="FF000000", size=16)

	for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
				cell.font = Font(color="FFFFFFFF", size=16)
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
				cell.font = Font(color="FF000000", size=16)
			else:
				if cell.value !=0:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
					cell.font = Font(color="FF000000", size=16)
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")
					cell.font = Font(color="FF000000", size=16)

	for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
		for cell in row:
			if ws1["I12"].value == "Korisnik preminuo":
				cell.fill = PatternFill("solid", fgColor="BA1E0E")
				cell.font = Font(color="FFFFFFFF", size=16)
			elif ws1["I12"].value == "Kredit otplaćen":
				cell.fill = PatternFill("solid", fgColor="A9F78D")
				cell.font = Font(color="FF000000", size=16)
			else:
				if cell.value !=None:
					cell.fill = PatternFill("solid", fgColor="A9F78D")
					cell.font = Font(color="FF000000", size=16)
				else:
					cell.fill = PatternFill("solid", fgColor="00FFCC99")
					cell.font = Font(color="FF000000", size=16)


def delete_user_excel(user_index):
	
	ws = wb["Zbirno"]

	target_row = user_index + 7

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
			sheet_name = str(row[1].value) + ". " + row[2].value			
				
	del wb[sheet_name]
				
	ws.delete_rows(target_row, 1)

	customer_names = []

	for row in ws.iter_rows(min_row=8):
		customer_names.append(row[2].value)

	customer_indexes = -1

	for sheet in wb.worksheets:
		customer_indexes+=1

	customer_indexes_list = list(range(1, customer_indexes+1))

	for index in (customer_indexes_list):
		ws["B" + str(index+7)] = index


	new_sheet_titles = ["Zbirno"]

	for index, name in zip(customer_indexes_list, customer_names):
		new_sheet_titles.append(str(index) + ". " + name)


	for sheet, new_sheet_name in zip(list(wb.worksheets), new_sheet_titles):
			sheet.title = new_sheet_name
			ws1 = wb[new_sheet_name]
			ws1['C1'] = new_sheet_name

def serial_payments_excel(user_index_list, today):

	payment_counter = 0

	for item in user_index_list:
		data, instalment_data, date_data, comment = confirm_selection_excel(item)
		if data[11] == "Otplata u toku":
			payment_counter += 1
			payment_excel(item, today)
		else:
			pass

	return payment_counter

def store_stats_excel(store_index):

	ws = wb["Zbirno"]

	users = []
	last_rows = []

	credits_total = 0
	total_value = 0

	closed_credits = 0
	paid_so_far = 0

	open_credits = 0
	open_credits_value = 0

	late_users = 0
	late_credits_value = 0

	all_store_credits = []

	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	if store_index.get() != "Sve MP":

		store = int(store_index.get())

		for row in ws.iter_rows(min_row=8):
			for cell in row:
				if cell.value == store_index.get():
					all_store_credits.append(row[6].value)
					sheet_name = str(row[1].value) + ". " + row[2].value
					instalment = row[7].value
					last_row = possible_cells[instalment+2]
					users.append(sheet_name)
					last_rows.append(last_row)
					credits_total +=1
	else:
		store = "Sve MP"

		for row in ws.iter_rows(min_row=8):
			sheet_name = str(row[1].value) + ". " + row[2].value
			instalment = row[7].value
			last_row = possible_cells[instalment+2]
			users.append(sheet_name)
			last_rows.append(last_row)
			credits_total +=1
	
	for user, last_row in zip(users, last_rows):

		ws1 = wb[user]

		credit = ws1["A17"].value
		total_value += credit

		balance = ws1[last_row + "17"].value
		status = ws1["I12"].value

		if status == "Otplata u toku":
			open_credits +=1
			open_credits_value += balance
			paid_portion = credit - balance
			paid_so_far += paid_portion
		
		if status == "Kredit otplaćen":
			closed_credits +=1
			paid_so_far += credit
		
		if status == "Korisnik preminuo":
			
			if balance < 1:
				paid_so_far += credit
				closed_credits +=1
			else:
				late_credits_value += balance
				paid_portion2 = credit - balance
				paid_so_far += paid_portion2

			late_users +=1

	if len(all_store_credits) == 0:
		max_credit_value = 0
		average_credit_value = 0
	else:
		max_credit_value = max(all_store_credits)
		average_credit_value = sum(all_store_credits)/len(all_store_credits)
		average_credit_value = round(average_credit_value, 2)

	total_remaining_value = late_credits_value + open_credits_value


	if total_value == 0:
		paid_percentage = 0

	else:
		paid_percentage = paid_so_far/total_value*100
		paid_percentage = round(paid_percentage, 2)


	total_value = round(total_value, 2)
	paid_so_far = round(paid_so_far, 2)
	open_credits_value = round(open_credits_value, 2)
	total_remaining_value = round(total_remaining_value, 2)
	late_credits_value = round(late_credits_value, 2)
	

	data = []

	data.extend([credits_total, total_value, closed_credits, paid_so_far, open_credits, open_credits_value, late_users, late_credits_value, total_remaining_value, max_credit_value, average_credit_value, paid_percentage, store])

	return data


def sort_stores_by_stats_excel(store_index):

	ws = wb["Zbirno"]

	store_index.set("51")
	MP51 = store_stats_excel(store_index)

	store_index.set("52")
	MP52 = store_stats_excel(store_index)

	store_index.set("53")
	MP53 = store_stats_excel(store_index)

	store_index.set("54")
	MP54 = store_stats_excel(store_index)

	store_index.set("55")
	MP55 = store_stats_excel(store_index)

	store_index.set("56")
	MP56 = store_stats_excel(store_index)

	store_index.set("59")
	MP59 = store_stats_excel(store_index)

	store_index.set("60")
	MP60 = store_stats_excel(store_index)

	store_index.set("61")
	MP61 = store_stats_excel(store_index)

	store_index.set("62")
	MP62 = store_stats_excel(store_index)

	store_index.set("63")
	MP63 = store_stats_excel(store_index)

	store_index.set("65")
	MP65 = store_stats_excel(store_index)

	store_index.set("66")
	MP66 = store_stats_excel(store_index)

	store_index.set("67")
	MP67 = store_stats_excel(store_index)

	store_index.set("68")
	MP68 = store_stats_excel(store_index)

	store_index.set("69")
	MP69 = store_stats_excel(store_index)

	store_index.set("70")
	MP70 = store_stats_excel(store_index)

	store_index.set("72")
	MP72 = store_stats_excel(store_index)

	store_index.set("73")
	MP73 = store_stats_excel(store_index)

	store_index.set("74")
	MP74 = store_stats_excel(store_index)

	data = []

	data.extend([MP51, MP52, MP53, MP54, MP55, MP56, MP59, MP60, MP61, MP62, MP63, MP65, MP66, MP67, MP68, MP69, MP70, MP72, MP73, MP74])

	return data

def add_payment_excel(unidentified_payment, unidentified_name, today):

	wb1 = load_workbook('unidentified_payments.xlsx')
	ws = wb1["Lista uplata"]

	row_counter = str(len(ws["E"])+1)

	index = len(ws['E']) - 6

	entry_data_cell = ["E", "F", "G", "H"]
	entry_data = [index, unidentified_name.get(), float(unidentified_payment.get()), today]
	for col, data in zip(entry_data_cell, entry_data):
		ws[col + row_counter] = data
		ws[col + row_counter].alignment = alignment_settings
		ws[col + row_counter].border = border_settings1

	ws["G" + row_counter].number_format = '#,##0.00'
		
	wb1.save('unidentified_payments.xlsx')

def populate_unidentified_search_box():

	wb1 = load_workbook('unidentified_payments.xlsx')
	ws = wb1["Lista uplata"]

	indexes = []
	names= []
	data = []

	for row in ws.iter_rows(min_col=5, max_col=5, min_row=8):
		for cell in row:
			indexes.append(cell.value)

	for row in ws.iter_rows(min_col=6, max_col=6, min_row=8):
		for cell in row:
			names.append(cell.value)

	for index, name in zip(indexes, names):
		display_name = str(index) + ". " + name
		data.append(display_name) 

	return data	

def select_unidentified_user_excel(value):

	wb1 = load_workbook('unidentified_payments.xlsx')
	ws = wb1["Lista uplata"]

	data = []

	index = re.findall(r'\d+', value)

	index = index = int(index[0])

	for row in ws.iter_rows(min_col=5, max_col=8, min_row=8):
		if row[0].value == index:
			data.extend([row[1].value, row[2].value, row[3].value])

	return data

def delete_unidentified_payment(value):

	wb1 = load_workbook('unidentified_payments.xlsx')
	ws = wb1["Lista uplata"]

	index = re.findall(r'\d+', value)

	index = int(index[0])

	target_row = index + 7

	ws.delete_rows(target_row, 1)
	
	number_of_items = 0

	for row in ws.iter_rows(min_col=5, max_col=5, min_row=8):
		for cell in row:
			if cell.value != None:
				number_of_items += 1

	new_index_numbers = list(range(1, number_of_items+1))

	for new_index in (new_index_numbers):
		ws["E" + str(new_index+7)] = new_index

	wb1.save('unidentified_payments.xlsx')

def payments_by_month_excel(drop):

	ws = wb["Zbirno"]

	users_with_payment = []

	users_without_payment_names = []
	users_without_payment_indexes = []
	
	possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

	pattern = fr"\d\d\.{drop.get()}\.+"

	if drop.get() == "Odaberi filter:":
		pass


	else:

		for row in ws.iter_rows(min_row=8):
			user_index = row[1].value
			sheet_name = str(row[1].value) + ". " + row[2].value
			instalment = row[7].value
			ws1 = wb[sheet_name]
			for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=instalment+1):
				for cell in row:
					if cell.value == None:
						pass

					else:
						match = re.match(pattern, cell.value)
						
						if match:
							users_with_payment.append(user_index)

						else:
							pass

		for row in ws.iter_rows(min_row=8):
			user_index = row[1].value
			if user_index in users_with_payment:
				pass
			else:
				name = row[2].value
				sheet_name = str(row[1].value) + ". " + row[2].value
				ws1 = wb[sheet_name]
				status = ws1["I12"].value
				if status == "Otplata u toku":
					users_without_payment_names.append(name)
					users_without_payment_indexes.append(user_index)
				else:
					pass


	return users_without_payment_names, users_without_payment_indexes


def print_user_excel(user_index):

	ws = wb["Zbirno"]

	wb.save('data.xlsx')

	file = "data.xlsx"


	my_path = os.path.abspath(os.path.dirname(file))
	file_path = os.path.join(my_path, "data.xlsx")


	
	o = win32com.client.Dispatch('Excel.Application')
	o.Visible = True
	wb2 = o.Workbooks.Open(file_path)
	ws2 = wb2.Worksheets([user_index + 1])
	ws2.PrintOut()
	wb2.Close(False)
	o.Quit()
	

def generate_store_workbook_excel(store_index):

	
	if store_index.get() != "Sve MP":

		ws = wb["Zbirno"]

		store_data = []

		possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

		for row in ws.iter_rows(min_row=8):
			if row[10].value == store_index.get():
				user_data = []
				data = []
				data.extend([row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value,row[9].value,row[10].value])
				sheet_name = str(row[1].value) + ". " + row[2].value
				table_header_length = row[7].value
				ws1 = wb[sheet_name]
				status = ws1["I12"].value
				comment = ws1["A22"].value
				data.append(status)
				last_row = possible_cells[table_header_length+2]
				cell_values = []
				cell_dates = []
				
				for row in ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=table_header_length+1):
					for cell in row:
						cell_values.append(cell.value)

				for row in ws1.iter_rows(min_row=18, max_row=18, min_col=2, max_col=table_header_length+1):
					for cell in row:
						cell_dates.append(cell.value)

				user_data.extend([data, cell_values, cell_dates])
				store_data.append(user_data)
				

		store_wb = Workbook()

		store_ws = store_wb.active

		store_ws.title = "Zbirno"

		sheet_ranges = store_wb["Zbirno"]
		sheet_ranges.column_dimensions["B"].width = 8
		sheet_ranges.column_dimensions["C"].width = 25
		sheet_ranges.column_dimensions["D"].width = 25
		sheet_ranges.column_dimensions["E"].width = 20
		sheet_ranges.column_dimensions["F"].width = 15
		sheet_ranges.column_dimensions["G"].width = 15
		sheet_ranges.column_dimensions["H"].width = 8
		sheet_ranges.column_dimensions["I"].width = 15
		sheet_ranges.column_dimensions["J"].width = 15
		sheet_ranges.column_dimensions["K"].width = 5


		store_ws['B2'] = f"PIO MALOPRODAJA {store_index.get()}"
		store_ws['B2'].font = Font(size=36, bold=True)
		store_ws['B2'].border = Border(left=Side(border_style='thick'),
											right=Side(border_style='thick'),
											top=Side(border_style='thick'),
											bottom=Side(border_style='thick'))
		store_ws['B2'].alignment = Alignment(horizontal='center', vertical= 'center')
				
		store_ws.merge_cells('B2:K3') 

		for col, column_name in zip(list(range(2,12)), ["Indeks", "Ime i prezime", "Adresa", "Grad", "JMBG", "Iznos kredita", "Br. rata", "Iznos rate", "Broj računa", "MP"]):
			c_letter = get_column_letter(col)  
			store_ws[c_letter + "7"] = column_name
			store_ws[c_letter + "7"].border = Border(left=Side(border_style='medium'),
								right=Side(border_style='medium'),
								top=Side(border_style='medium'),
								bottom=Side(border_style='medium'))
			store_ws[c_letter + "7"].alignment = Alignment(horizontal='center', vertical= 'center')



		entry_data_cell = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

		ws_index = 1
		ws_row_counter = 8
		
		for item in store_data:
			entry_data = item[0]
			entry_data.insert(0, ws_index)
			value_data = item[1]
			date_data = item[2]

			new_sheet_name = str(ws_index) + ". " + entry_data[1]
			store_wb.create_sheet(title= new_sheet_name)

			instalments = entry_data[6]

			store_ws1 = store_wb[new_sheet_name]
			sheet_ranges = store_wb[new_sheet_name]

			sheet_ranges.column_dimensions["A"].width = 15
			sheet_ranges.column_dimensions["B"].width = 15
			sheet_ranges.column_dimensions["C"].width = 15
			sheet_ranges.column_dimensions["D"].width = 15
			sheet_ranges.column_dimensions["E"].width = 15
			sheet_ranges.column_dimensions["F"].width = 15
			sheet_ranges.column_dimensions["G"].width = 15
			sheet_ranges.column_dimensions["H"].width = 15
			sheet_ranges.column_dimensions["I"].width = 15
			sheet_ranges.column_dimensions["J"].width = 15
			sheet_ranges.column_dimensions["K"].width = 15
			sheet_ranges.column_dimensions["L"].width = 15
			sheet_ranges.column_dimensions["M"].width = 15
			sheet_ranges.column_dimensions["N"].width = 15

			sheet_ranges.row_dimensions[16].height = 63
			sheet_ranges.row_dimensions[17].height = 63
			sheet_ranges.row_dimensions[18].height = 63

		    
			store_ws1.sheet_view.zoomScale = 70
			

			store_ws1['C1'] = new_sheet_name
			store_ws1['C1'].font = Font(size=36, bold=True)
			store_ws1['C1'].border = border_settings2
			store_ws1['C1'].alignment = alignment_settings
			store_ws1.merge_cells('C1:L3')


			store_ws1['L6'] = entry_data[9]
			store_ws1['L6'].font = Font(size=36, bold=True)
			store_ws1['L6'].border = border_settings3
			store_ws1['L6'].alignment = alignment_settings
			store_ws1.merge_cells('L6:L9')

			store_ws1['C6'] = "Grad"
			store_ws1['C6'].font = Font(size=16)
			store_ws1['C6'].border = border_settings3
			store_ws1['C6'].alignment = alignment_settings
			store_ws1.merge_cells('C6:D7')


			store_ws1['C8'] = "Adresa"
			store_ws1['C8'].font = Font(size=16)
			store_ws1['C8'].border = border_settings3
			store_ws1['C8'].alignment = alignment_settings
			store_ws1.merge_cells('C8:D9')


			store_ws1['C10'] = "JMBG"
			store_ws1['C10'].font = Font(size=16)
			store_ws1['C10'].border = border_settings3
			store_ws1['C10'].alignment = alignment_settings
			store_ws1.merge_cells('C10:D11')	


			store_ws1['C12'] = "Broj računa"
			store_ws1['C12'].font = Font(size=16)
			store_ws1['C12'].border = border_settings3
			store_ws1['C12'].alignment = alignment_settings
			store_ws1.merge_cells('C12:D13')


			store_ws1['E6'] = entry_data[3]
			store_ws1['E6'].font = Font(size=16)
			store_ws1['E6'].border = border_settings3
			store_ws1['E6'].alignment = alignment_settings
			store_ws1.merge_cells('E6:G7')


			store_ws1['E8'] = entry_data[2]
			store_ws1['E8'].font = Font(size=16)
			store_ws1['E8'].border = border_settings3
			store_ws1['E8'].alignment = alignment_settings
			store_ws1.merge_cells('E8:G9')


			store_ws1['E10'] = entry_data[4]
			store_ws1['E10'].font = Font(size=16)
			store_ws1['E10'].border = border_settings3
			store_ws1['E10'].alignment = alignment_settings
			store_ws1.merge_cells('E10:G11')


			store_ws1['E12'] = entry_data[8]
			store_ws1['E12'].font = Font(size=16)
			store_ws1['E12'].border = border_settings3
			store_ws1['E12'].alignment = alignment_settings
			store_ws1.merge_cells('E12:G13')

			store_ws1['I12'] = entry_data[10]
			store_ws1['I12'].font = Font(size=16)
			store_ws1['I12'].border = border_settings3
			store_ws1['I12'].alignment = alignment_settings

			store_ws1['A20'] = "Napomene"
			store_ws1['A20'].font = Font(size=26)
			store_ws1['A20'].border = border_settings3
			store_ws1['A20'].alignment = alignment_settings
			store_ws1.merge_cells('A20:N21')

			
			store_ws1['A22'].border = border_settings3
			store_ws1['A22'].value = comment
			store_ws1['A22'].font = Font(size=16)
			store_ws1['A22'].alignment = Alignment(horizontal="left", vertical="top")
			store_ws1.merge_cells('A22:N30')

			if entry_data[10] == "Korisnik preminuo":
				store_ws1['I12'].fill = PatternFill("solid", fgColor="BA1E0E")
			elif entry_data[10] == "Kredit otplaćen":
				store_ws1['I12'].fill = PatternFill("solid", fgColor="A9F78D")
			else:
				store_ws1['I12'].fill = PatternFill("solid", fgColor="00FFCC99")
			store_ws1.merge_cells('I12:L13')

			header_list = ["Početno"]

			instalments_list = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]

			end_list = "Ostatak"

			for no in range(int(instalments)):
				header_list.append(instalments_list[no])
		    
			header_list.append(end_list)
			
			for col, column_name in zip(list(range(1, len(header_list)+1)), header_list):
				c_letter = get_column_letter(col)  
				store_ws1[c_letter + "16"] = column_name
				store_ws1[c_letter + "16"].border = border_settings6
				store_ws1[c_letter + "16"].alignment = alignment_settings
				store_ws1[c_letter + "16"].font = Font(size=20, bold=True)

			for col in range(2,len(header_list)+1):
				c_letter = get_column_letter(col)
				store_ws1[c_letter + "17"].border = border_settings1
				store_ws1[c_letter + "17"].alignment = alignment_settings
				store_ws1[c_letter + "17"].font = Font(size=16)
				store_ws1[c_letter + "17"].number_format = '#,##0.00'

			for col, value in zip(range(2,len(header_list)), value_data):
				c_letter = get_column_letter(col)
				store_ws1[c_letter + "17"].value = value
				if entry_data[10] == "Korisnik preminuo":
					store_ws1[c_letter + "17"].fill = PatternFill("solid", fgColor="BA1E0E")
				else:
					if value == 0:
						store_ws1[c_letter + "17"].fill = PatternFill("solid", fgColor="00FFCC99")
					else:
						store_ws1[c_letter + "17"].fill = PatternFill("solid", fgColor="A9F78D")

			for col in range(1,len(header_list)+1):
				c_letter = get_column_letter(col)
				store_ws1[c_letter + "18"].border = border_settings7
				store_ws1[c_letter + "18"].alignment = alignment_settings
				store_ws1[c_letter + "18"].font = Font(size=16)

			for col, date in zip(range(2,len(header_list)), date_data):
				c_letter = get_column_letter(col)
				store_ws1[c_letter + "18"].value = date
				if entry_data[10] == "Korisnik preminuo":
					store_ws1[c_letter + "18"].fill = PatternFill("solid", fgColor="BA1E0E")
				else:
					if date == None:
						store_ws1[c_letter + "18"].fill = PatternFill("solid", fgColor="00FFCC99")
					else:
						store_ws1[c_letter + "18"].fill = PatternFill("solid", fgColor="A9F78D")

			possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

			last_row = possible_cells[len(header_list)]

			store_ws1['A16'].border = Border(left=Side(border_style='medium'),
									right=Side(border_style='thin'),
									top=Side(border_style='medium'),
									bottom=Side(border_style='thin'))
			store_ws1['A16'].alignment = alignment_settings	
			
			store_ws1["A17"] = float(entry_data[5])
			store_ws1["A17"].number_format = '#,##0.00'

			store_ws1['A17'].border = border_settings4
			store_ws1['A17'].alignment = alignment_settings
			store_ws1['A17'].font = Font(size=16)		
				

			store_ws1["A18"] = "Datum Uplate"
			store_ws1['A18'].border = Border(left=Side(border_style='medium'),
									right=Side(border_style='thin'),
									top=Side(border_style='thin'),
									bottom=Side(border_style='medium'))
			store_ws1['A18'].alignment = alignment_settings
			store_ws1['A18'].font = Font(size=12, bold=True)

			
			store_ws1[last_row + '16'].border = Border(left=Side(border_style='thin'),
									right=Side(border_style='medium'),
									top=Side(border_style='medium'),
									bottom=Side(border_style='thin'))
			store_ws1[last_row + '16'].alignment = alignment_settings	


			store_ws1[last_row + '17'].border = border_settings5
			store_ws1[last_row + '17'].alignment = alignment_settings

			middle_rows = []

			for row in store_ws1.iter_rows(min_row=17, max_row=17, min_col=2, max_col=len(header_list)-1):
				for cell in row:
					cell_name = cell.coordinate
					cell_value = store_ws1[cell_name].value
					middle_rows.append(cell_value)

			
			store_ws1[ last_row + "17"].number_format = '#,##0.00'
			store_ws1[ last_row + "17"] = store_ws1["A17"].value - sum(middle_rows)
			store_ws1[ last_row + "17"].value = round(float(store_ws1[ last_row + "17"].value),2)	
			
			store_ws1[last_row + '18'].border = Border(left=Side(border_style='thin'),
									right=Side(border_style='medium'),
									top=Side(border_style='thin'),
									bottom=Side(border_style='medium'))
			store_ws1[last_row + '18'].alignment = alignment_settings

			for col, data in zip(entry_data_cell, entry_data):
				store_ws[col + str(ws_row_counter)] = data
				store_ws[col + str(ws_row_counter)].alignment = alignment_settings
				store_ws[col + str(ws_row_counter)].border = border_settings1
				store_ws["G" + str(ws_row_counter)].number_format = '#,##0.00'
				store_ws["I" + str(ws_row_counter)].number_format = '#,##0.00'
				if entry_data[10] == "Korisnik preminuo":
					store_ws[col + str(ws_row_counter)].fill = PatternFill("solid", fgColor="BA1E0E")
				elif entry_data[10] == "Kredit otplaćen":
					store_ws[col + str(ws_row_counter)].fill = PatternFill("solid", fgColor="A9F78D")
				else:
					store_ws[col + str(ws_row_counter)].fill = PatternFill("solid", fgColor="00FFCC99")

			ws_index +=1
			ws_row_counter +=1

		store_stat_data = store_stats_excel(store_index)

		total_value = store_stat_data[1] 
		closed_credits_value = store_stat_data[3]
		open_credits_value = store_stat_data[5]
		late_credits_value = store_stat_data[7]
		total_remaining_value = store_stat_data[8]

		total_value = str(f'{total_value:,}') + " din."
		closed_credits_value = str(f'{closed_credits_value:,}') + " din."
		open_credits_value = str(f'{open_credits_value:,}') + " din."
		total_remaining_value = str(f'{total_remaining_value:,}') + " din."
		late_credits_value = str(f'{late_credits_value:,}') + " din."


		store_ws['D' + str(ws_row_counter+3)] = "Ukupan iznos"
		store_ws['D' + str(ws_row_counter+3)].font = Font(size=16)
		store_ws['D' + str(ws_row_counter+3)].border = border_settings3
		store_ws['D' + str(ws_row_counter+3)].alignment = alignment_settings
		store_ws.merge_cells('D' + str(ws_row_counter+3) + ':' + 'D' + str(ws_row_counter+4))

		store_ws['D' + str(ws_row_counter+5)] = "Isplaćeno"
		store_ws['D' + str(ws_row_counter+5)].font = Font(size=16)
		store_ws['D' + str(ws_row_counter+5)].border = border_settings3
		store_ws['D' + str(ws_row_counter+5)].alignment = alignment_settings
		store_ws['D' + str(ws_row_counter+5)].fill = PatternFill("solid", fgColor="A9F78D")
		store_ws.merge_cells('D' + str(ws_row_counter+5) + ':' + 'D' + str(ws_row_counter+6))

		store_ws['D' + str(ws_row_counter+7)] = "Otplata u toku"
		store_ws['D' + str(ws_row_counter+7)].font = Font(size=16)
		store_ws['D' + str(ws_row_counter+7)].border = border_settings3
		store_ws['D' + str(ws_row_counter+7)].alignment = alignment_settings
		store_ws['D' + str(ws_row_counter+7)].fill = PatternFill("solid", fgColor="00FFCC99")
		store_ws.merge_cells('D' + str(ws_row_counter+7) + ':' + 'D' + str(ws_row_counter+8))

		store_ws['D' + str(ws_row_counter+9)] = "Preminuli"
		store_ws['D' + str(ws_row_counter+9)].font = Font(size=16)
		store_ws['D' + str(ws_row_counter+9)].border = border_settings3
		store_ws['D' + str(ws_row_counter+9)].alignment = alignment_settings
		store_ws['D' + str(ws_row_counter+9)].fill = PatternFill("solid", fgColor="BA1E0E")
		store_ws.merge_cells('D' + str(ws_row_counter+9) + ':' + 'D' + str(ws_row_counter+10))

		store_ws['D' + str(ws_row_counter+11)] = "Ukupno duga"
		store_ws['D' + str(ws_row_counter+11)].font = Font(size=16)
		store_ws['D' + str(ws_row_counter+11)].border = border_settings3
		store_ws['D' + str(ws_row_counter+11)].alignment = alignment_settings
		store_ws.merge_cells('D' + str(ws_row_counter+11) + ':' + 'D' + str(ws_row_counter+12))

		store_ws['E' + str(ws_row_counter+3)] = total_value
		store_ws['E' + str(ws_row_counter+3)].font = Font(size=16)
		store_ws['E' + str(ws_row_counter+3)].border = border_settings3
		store_ws['E' + str(ws_row_counter+3)].alignment = alignment_settings
		store_ws.merge_cells('E' + str(ws_row_counter+3) + ':' + 'F' + str(ws_row_counter+4))

		store_ws['E' + str(ws_row_counter+5)] =closed_credits_value
		store_ws['E' + str(ws_row_counter+5)].font = Font(size=16)
		store_ws['E' + str(ws_row_counter+5)].border = border_settings3
		store_ws['E' + str(ws_row_counter+5)].alignment = alignment_settings
		store_ws['E' + str(ws_row_counter+5)].fill = PatternFill("solid", fgColor="A9F78D")
		store_ws.merge_cells('E' + str(ws_row_counter+5) + ':' + 'F' + str(ws_row_counter+6))

		store_ws['E' + str(ws_row_counter+7)] =open_credits_value
		store_ws['E' + str(ws_row_counter+7)].font = Font(size=16)
		store_ws['E' + str(ws_row_counter+7)].border = border_settings3
		store_ws['E' + str(ws_row_counter+7)].alignment = alignment_settings
		store_ws['E' + str(ws_row_counter+7)].fill = PatternFill("solid", fgColor="00FFCC99")
		store_ws.merge_cells('E' + str(ws_row_counter+7) + ':' + 'F' + str(ws_row_counter+8))

		store_ws['E' + str(ws_row_counter+9)] =late_credits_value
		store_ws['E' + str(ws_row_counter+9)].font = Font(size=16)
		store_ws['E' + str(ws_row_counter+9)].border = border_settings3
		store_ws['E' + str(ws_row_counter+9)].alignment = alignment_settings
		store_ws['E' + str(ws_row_counter+9)].fill = PatternFill("solid", fgColor="BA1E0E")
		store_ws.merge_cells('E' + str(ws_row_counter+9) + ':' + 'F' + str(ws_row_counter+10))

		store_ws['E' + str(ws_row_counter+11)] =total_remaining_value
		store_ws['E' + str(ws_row_counter+11)].font = Font(size=16)
		store_ws['E' + str(ws_row_counter+11)].border = border_settings3
		store_ws['E' + str(ws_row_counter+11)].alignment = alignment_settings
		store_ws.merge_cells('E' + str(ws_row_counter+11) + ':' + 'F' + str(ws_row_counter+12))
		
		store_wb.save(f'Maloprodaja {store_index.get()}.xlsx')

	else:
		wb.save('data.xlsx')

		status_wb = load_workbook('data.xlsx')

		status_ws = status_wb["Zbirno"]

		store_data = []

		possible_cells = [0, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

		for row in status_ws.iter_rows(min_row=8):
			sheet_name = str(row[1].value) + ". " + row[2].value
			table_header_length = row[7].value
			status_ws1 = status_wb[sheet_name]
			status = status_ws1["I12"].value
			if status == "Korisnik preminuo":
				for cell in row:
					if cell.value != None:
						cell.fill = PatternFill("solid", fgColor="BA1E0E")
			elif status == "Kredit otplaćen":
				for cell in row:
					if cell.value != None:
						cell.fill = PatternFill("solid", fgColor="A9F78D")
			else:
				for cell in row:
					if cell.value != None:
						cell.fill = PatternFill("solid", fgColor="00FFCC99")

		status_row_counter = len(status_ws["B"])+1

		store_stat_data = store_stats_excel(store_index)

		total_value = store_stat_data[1] 
		closed_credits_value = store_stat_data[3]
		open_credits_value = store_stat_data[5]
		late_credits_value = store_stat_data[7]
		total_remaining_value = store_stat_data[8]

		total_value = str(f'{total_value:,}') + " din."
		closed_credits_value = str(f'{closed_credits_value:,}') + " din."
		open_credits_value = str(f'{open_credits_value:,}') + " din."
		total_remaining_value = str(f'{total_remaining_value:,}') + " din."
		late_credits_value = str(f'{late_credits_value:,}') + " din."


		status_ws['D' + str(status_row_counter+3)] = "Ukupan iznos"
		status_ws['D' + str(status_row_counter+3)].font = Font(size=16)
		status_ws['D' + str(status_row_counter+3)].border = border_settings3
		status_ws['D' + str(status_row_counter+3)].alignment = alignment_settings
		status_ws.merge_cells('D' + str(status_row_counter+3) + ':' + 'D' + str(status_row_counter+4))

		status_ws['D' + str(status_row_counter+5)] = "Isplaćeno"
		status_ws['D' + str(status_row_counter+5)].font = Font(size=16)
		status_ws['D' + str(status_row_counter+5)].border = border_settings3
		status_ws['D' + str(status_row_counter+5)].alignment = alignment_settings
		status_ws['D' + str(status_row_counter+5)].fill = PatternFill("solid", fgColor="A9F78D")
		status_ws.merge_cells('D' + str(status_row_counter+5) + ':' + 'D' + str(status_row_counter+6))

		status_ws['D' + str(status_row_counter+7)] = "Otplata u toku"
		status_ws['D' + str(status_row_counter+7)].font = Font(size=16)
		status_ws['D' + str(status_row_counter+7)].border = border_settings3
		status_ws['D' + str(status_row_counter+7)].alignment = alignment_settings
		status_ws['D' + str(status_row_counter+7)].fill = PatternFill("solid", fgColor="00FFCC99")
		status_ws.merge_cells('D' + str(status_row_counter+7) + ':' + 'D' + str(status_row_counter+8))

		status_ws['D' + str(status_row_counter+9)] = "Preminuli"
		status_ws['D' + str(status_row_counter+9)].font = Font(size=16)
		status_ws['D' + str(status_row_counter+9)].border = border_settings3
		status_ws['D' + str(status_row_counter+9)].alignment = alignment_settings
		status_ws['D' + str(status_row_counter+9)].fill = PatternFill("solid", fgColor="BA1E0E")
		status_ws.merge_cells('D' + str(status_row_counter+9) + ':' + 'D' + str(status_row_counter+10))

		status_ws['D' + str(status_row_counter+11)] = "Ukupno duga"
		status_ws['D' + str(status_row_counter+11)].font = Font(size=16)
		status_ws['D' + str(status_row_counter+11)].border = border_settings3
		status_ws['D' + str(status_row_counter+11)].alignment = alignment_settings
		status_ws.merge_cells('D' + str(status_row_counter+11) + ':' + 'D' + str(status_row_counter+12))

		status_ws['E' + str(status_row_counter+3)] = total_value
		status_ws['E' + str(status_row_counter+3)].font = Font(size=16)
		status_ws['E' + str(status_row_counter+3)].border = border_settings3
		status_ws['E' + str(status_row_counter+3)].alignment = alignment_settings
		status_ws.merge_cells('E' + str(status_row_counter+3) + ':' + 'F' + str(status_row_counter+4))

		status_ws['E' + str(status_row_counter+5)] =closed_credits_value
		status_ws['E' + str(status_row_counter+5)].font = Font(size=16)
		status_ws['E' + str(status_row_counter+5)].border = border_settings3
		status_ws['E' + str(status_row_counter+5)].alignment = alignment_settings
		status_ws['E' + str(status_row_counter+5)].fill = PatternFill("solid", fgColor="A9F78D")
		status_ws.merge_cells('E' + str(status_row_counter+5) + ':' + 'F' + str(status_row_counter+6))

		status_ws['E' + str(status_row_counter+7)] =open_credits_value
		status_ws['E' + str(status_row_counter+7)].font = Font(size=16)
		status_ws['E' + str(status_row_counter+7)].border = border_settings3
		status_ws['E' + str(status_row_counter+7)].alignment = alignment_settings
		status_ws['E' + str(status_row_counter+7)].fill = PatternFill("solid", fgColor="00FFCC99")
		status_ws.merge_cells('E' + str(status_row_counter+7) + ':' + 'F' + str(status_row_counter+8))

		status_ws['E' + str(status_row_counter+9)] =late_credits_value
		status_ws['E' + str(status_row_counter+9)].font = Font(size=16)
		status_ws['E' + str(status_row_counter+9)].border = border_settings3
		status_ws['E' + str(status_row_counter+9)].alignment = alignment_settings
		status_ws['E' + str(status_row_counter+9)].fill = PatternFill("solid", fgColor="BA1E0E")
		status_ws.merge_cells('E' + str(status_row_counter+9) + ':' + 'F' + str(status_row_counter+10))

		status_ws['E' + str(status_row_counter+11)] =total_remaining_value
		status_ws['E' + str(status_row_counter+11)].font = Font(size=16)
		status_ws['E' + str(status_row_counter+11)].border = border_settings3
		status_ws['E' + str(status_row_counter+11)].alignment = alignment_settings
		status_ws.merge_cells('E' + str(status_row_counter+11) + ':' + 'F' + str(status_row_counter+12))

		status_wb.save('Presek stanja - SVE MP.xlsx')


def write_comment_excel(user_comment, user_index):

	ws = wb["Zbirno"]

	for row in ws.iter_rows(min_row=8):
		if row[1].value == user_index:
			sheet_name = str(row[1].value) + ". " + row[2].value

	ws1 = wb[sheet_name]

	ws1['A22'] = user_comment

	




	


	


